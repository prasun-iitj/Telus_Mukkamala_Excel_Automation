"""Transformer: converts raw Excel data into structured DataModel."""

import logging
from typing import Optional

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from app.models import (
    CellValue,
    DataModel,
    MonthData,
    SectionData,
    TeamMetrics,
)
from app.readers.excel_reader import ExcelReader

logger = logging.getLogger(__name__)

# Section headers to scan for
SECTION_HEADERS = {
    "delivery": "On-Time and Planned Delivery",
    "quality": "Quality and Dev Efficiency",
    "rollout": "Rollout Success",
    "product_health": "Product Health Measure",
}

# Month ordering for auto-detection
MONTH_ORDER = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


class Transformer:
    """Converts raw Excel cell data into structured DataModel."""

    def extract_section(
        self,
        reader: ExcelReader,
        sheet_name: str,
        section_header: str,
    ) -> SectionData:
        """Find a section by header text and extract all month blocks.

        Scans column B for the section header, reads column headers from the
        next row, then iterates over month blocks using merged-cell ranges in
        column B to determine boundaries.

        Args:
            reader: An initialised ExcelReader.
            sheet_name: Name of the worksheet.
            section_header: The header text to search for (e.g. "Rollout Success").

        Returns:
            A SectionData containing all months found in the section.
        """
        sheet = reader.get_sheet(sheet_name)

        # --- locate section header row ---
        header_row = self._find_section_header_row(reader, sheet, section_header)
        if header_row is None:
            logger.warning(
                "Section header '%s' not found in sheet '%s'",
                section_header,
                sheet_name,
            )
            return SectionData(
                section_name=section_header,
                sheet_name=sheet_name,
                months=[],
                column_headers=[],
            )

        # --- read column headers (row immediately after section header) ---
        col_header_row = header_row + 1
        column_headers = self._read_column_headers(reader, sheet, col_header_row)

        # --- determine which columns are percentage columns ---
        pct_col_indices = self._detect_percentage_columns(column_headers)

        # --- determine section type for end-of-section detection ---
        section_type = self._section_type_for(section_header)

        # --- collect month blocks via merged ranges in column B ---
        month_blocks = self._collect_month_blocks(sheet, header_row)

        # --- extract data for each month block ---
        months: list[MonthData] = []
        max_data_col = self._detect_max_data_col(column_headers)

        for month_name, start_row, end_row in month_blocks:
            month_data = self.extract_month_block(
                reader,
                sheet,
                start_row,
                end_row,
                column_headers,
                pct_col_indices,
                max_data_col,
            )
            # Override month name from merged cell (authoritative source)
            month_data.month_name = month_name
            months.append(month_data)

        return SectionData(
            section_name=section_header,
            sheet_name=sheet_name,
            months=months,
            column_headers=column_headers,
        )

    def extract_month_block(
        self,
        reader: ExcelReader,
        sheet: Worksheet,
        start_row: int,
        end_row: int,
        column_headers: list[str],
        pct_col_indices: set[int],
        max_data_col: int,
    ) -> MonthData:
        """Extract one month's data across all teams.

        Args:
            reader: An initialised ExcelReader.
            sheet: The worksheet object.
            start_row: First data row of the month block (inclusive).
            end_row: Last data row of the month block (inclusive).
            column_headers: List of metric column names (0-indexed, starting from col D).
            pct_col_indices: Set of 0-based indices into column_headers that are percentage columns.
            max_data_col: 1-based max column to read.

        Returns:
            A MonthData with dynamically extracted team names and metrics.
        """
        teams: list[TeamMetrics] = []

        for row in range(start_row, end_row + 1):
            # Read team name from column C (col 3)
            team_name = reader.read_cell(sheet, row, 3)
            if team_name is None or (isinstance(team_name, str) and team_name.strip() == ""):
                continue  # skip empty rows

            team_name = str(team_name).strip()

            # Read metric values from col D (4) onwards
            values: dict[str, CellValue] = {}
            for i, header in enumerate(column_headers):
                col = 4 + i  # col D = 4
                if col > max_data_col:
                    break
                raw_value = reader.read_cell(sheet, row, col)
                # Convert decimal percentages to percentage representation
                if i in pct_col_indices and isinstance(raw_value, (int, float)):
                    if 0.0 <= raw_value <= 1.0:
                        raw_value = round(raw_value * 100, 4)
                values[header] = raw_value

            teams.append(TeamMetrics(team_name=team_name, values=values))

        # Month name will be overridden by caller from merged cell
        return MonthData(month_name="", teams=teams)

    def build_data_model(self, reader: ExcelReader) -> DataModel:
        """Build the complete DataModel from both Master sheets.

        Extracts all three sections (delivery, quality, rollout) from both
        Master-Apps and Master-Platforms sheets.

        Args:
            reader: An initialised ExcelReader.

        Returns:
            A fully populated DataModel.
        """
        apps_sheet = "Master-Apps (DoNotChange)"
        platforms_sheet = "Master-Platforms (DoNotChange)"

        apps_delivery = self.extract_section(
            reader, apps_sheet, SECTION_HEADERS["delivery"]
        )
        apps_quality = self.extract_section(
            reader, apps_sheet, SECTION_HEADERS["quality"]
        )
        apps_rollout = self.extract_section(
            reader, apps_sheet, SECTION_HEADERS["rollout"]
        )
        platforms_delivery = self.extract_section(
            reader, platforms_sheet, SECTION_HEADERS["delivery"]
        )
        platforms_quality = self.extract_section(
            reader, platforms_sheet, SECTION_HEADERS["quality"]
        )
        platforms_rollout = self.extract_section(
            reader, platforms_sheet, SECTION_HEADERS["rollout"]
        )
        apps_product_health = self.extract_section(
            reader, apps_sheet, SECTION_HEADERS["product_health"]
        )
        platforms_product_health = self.extract_section(
            reader, platforms_sheet, SECTION_HEADERS["product_health"]
        )

        data_model = DataModel(
            apps_delivery=apps_delivery,
            apps_quality=apps_quality,
            apps_rollout=apps_rollout,
            apps_product_health=apps_product_health,
            platforms_delivery=platforms_delivery,
            platforms_quality=platforms_quality,
            platforms_rollout=platforms_rollout,
            platforms_product_health=platforms_product_health,
            reporting_month="",
        )

        # Auto-detect reporting month
        data_model.reporting_month = self.detect_reporting_month(data_model)

        return data_model

    def detect_reporting_month(self, data_model: DataModel) -> str:
        """Find the latest month with real reporting data across multiple sections.

        A month qualifies only when it has non-YTD numeric data in at least
        ``MIN_SECTIONS_FOR_MONTH`` of the eight Master-sheet sections.  This
        avoids false positives from stray YTD values or template placeholders
        in future month blocks (e.g. April detected when March is intended).
        """
        MIN_SECTIONS_FOR_MONTH = 3

        all_sections = [
            data_model.apps_delivery,
            data_model.apps_quality,
            data_model.apps_rollout,
            data_model.apps_product_health,
            data_model.platforms_delivery,
            data_model.platforms_quality,
            data_model.platforms_rollout,
            data_model.platforms_product_health,
        ]

        month_section_counts: dict[str, int] = {}
        for section in all_sections:
            for month in section.months:
                if self._month_has_data(month):
                    month_section_counts[month.month_name] = (
                        month_section_counts.get(month.month_name, 0) + 1
                    )

        qualified = {
            name for name, count in month_section_counts.items()
            if count >= MIN_SECTIONS_FOR_MONTH
        }
        if not qualified:
            # Fallback: any month with data in at least one section
            qualified = set(month_section_counts.keys())
        if not qualified:
            return ""

        latest_idx = -1
        latest_name = ""
        for name in qualified:
            try:
                idx = MONTH_ORDER.index(name)
            except ValueError:
                continue
            if idx > latest_idx:
                latest_idx = idx
                latest_name = name

        return latest_name

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _find_section_header_row(
        reader: ExcelReader,
        sheet: Worksheet,
        section_header: str,
    ) -> Optional[int]:
        """Scan column B for a row whose value matches *section_header*."""
        for row in range(1, sheet.max_row + 1):
            value = reader.read_cell(sheet, row, 2)  # column B
            if value is not None and str(value).strip() == section_header:
                return row
        return None

    @staticmethod
    def _read_column_headers(
        reader: ExcelReader,
        sheet: Worksheet,
        row: int,
    ) -> list[str]:
        """Read column headers starting from col D until values run out."""
        headers: list[str] = []
        col = 4  # col D
        while col <= sheet.max_column:
            val = reader.read_cell(sheet, row, col)
            if val is None:
                break
            # Normalise non-breaking spaces and strip
            header_str = str(val).replace("\xa0", " ").strip()
            headers.append(header_str)
            col += 1
        return headers

    @staticmethod
    def _detect_percentage_columns(column_headers: list[str]) -> set[int]:
        """Return 0-based indices of columns whose header contains '%' or 'Percentage'."""
        pct_indices: set[int] = set()
        for i, header in enumerate(column_headers):
            lower = header.lower()
            if "%" in header or "percentage" in lower:
                pct_indices.add(i)
        return pct_indices

    @staticmethod
    def _detect_max_data_col(column_headers: list[str]) -> int:
        """Return the 1-based max column index for data (col D + len(headers) - 1)."""
        return 4 + len(column_headers) - 1 if column_headers else 4

    @staticmethod
    def _collect_month_blocks(
        sheet: Worksheet,
        section_header_row: int,
    ) -> list[tuple[str, int, int]]:
        """Collect (month_name, start_row, end_row) tuples from merged ranges in col B.

        Only considers merged ranges that start after the column-header row
        (section_header_row + 1) and whose value is a recognised month name.
        Stops after collecting all 12 months or when a gap indicates a new section.
        """
        col_header_row = section_header_row + 1
        blocks: list[tuple[str, int, int]] = []

        # Gather all merged ranges in column B that are below the header row,
        # sorted by row position
        candidates = []
        for mr in sheet.merged_cells.ranges:
            if mr.min_col != 2 or mr.max_col != 2:
                continue
            if mr.min_row <= col_header_row:
                continue
            candidates.append(mr)
        candidates.sort(key=lambda r: r.min_row)

        last_end_row = col_header_row
        for mr in candidates:
            # If there's a large gap (more than 2 rows) between the last block
            # and this one, we've likely hit a new section
            if blocks and (mr.min_row - last_end_row) > 2:
                break

            month_val = sheet.cell(row=mr.min_row, column=mr.min_col).value
            if month_val is None:
                continue
            month_name = str(month_val).strip()
            if month_name in MONTH_ORDER:
                blocks.append((month_name, mr.min_row, mr.max_row))
                last_end_row = mr.max_row
            else:
                # Not a month name — we've hit the next section
                break

            # Stop after 12 months (one full year)
            if len(blocks) >= 12:
                break

        return blocks

    @staticmethod
    def _is_ytd_metric(metric_name: str) -> bool:
        """Return True for year-to-date columns that should not drive month detection."""
        lower = metric_name.lower()
        return "(ytd)" in lower or "ytd)" in lower or lower.endswith("(ytd)")

    @staticmethod
    def _month_has_data(month: MonthData) -> bool:
        """Return True if at least one team has a non-null, non-zero monthly metric.

        YTD columns are excluded — they often carry cumulative values in empty
        future month blocks and cause false reporting-month detection.
        """
        for team in month.teams:
            for key, value in team.values.items():
                if Transformer._is_ytd_metric(key):
                    continue
                if value is None:
                    continue
                if isinstance(value, (int, float)) and value != 0:
                    return True
        return False

    @staticmethod
    def _section_type_for(section_header: str) -> str:
        """Map a section header to a short type key."""
        lower = section_header.lower()
        if "delivery" in lower:
            return "delivery"
        if "quality" in lower:
            return "quality"
        if "rollout" in lower:
            return "rollout"
        return "unknown"
