"""Excel workbook reader with merged cell handling."""

import logging
import os
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.worksheet.worksheet import Worksheet

from app.exceptions import PipelineError, SheetNotFoundError
from app.models import CellValue

logger = logging.getLogger(__name__)


class ExcelReader:
    """Loads an Excel workbook and provides cell/range reading with merged cell support."""

    def __init__(self, file_path: str) -> None:
        """Load workbook with data_only=True to get computed values.

        Args:
            file_path: Path to the .xlsx file.

        Raises:
            PipelineError: If the path is invalid or not a .xlsx file.
        """
        if not file_path.endswith(".xlsx"):
            raise PipelineError("excel_reader", f"Invalid file path: '{file_path}' is not a .xlsx file")
        if not os.path.isfile(file_path):
            raise PipelineError("excel_reader", f"Invalid file path: '{file_path}' does not exist")

        self.file_path = file_path
        self.workbook = load_workbook(file_path, data_only=True)

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """Return worksheet by name.

        Args:
            sheet_name: Name of the sheet to retrieve.

        Returns:
            The requested worksheet.

        Raises:
            SheetNotFoundError: If the sheet does not exist.
        """
        if sheet_name not in self.workbook.sheetnames:
            raise SheetNotFoundError(sheet_name)
        return self.workbook[sheet_name]

    def read_cell(self, sheet: Worksheet, row: int, col: int) -> CellValue:
        """Read a single cell, resolving merged cell ranges to top-left value.

        Args:
            sheet: The worksheet to read from.
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            The cell value, or the top-left value if the cell is merged.
        """
        cell = sheet.cell(row=row, column=col)

        # If the cell is a MergedCell, find the top-left cell of the merged range
        if isinstance(cell, MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_row = merged_range.min_row
                    top_col = merged_range.min_col
                    value = sheet.cell(row=top_row, column=top_col).value
                    return self._process_value(value, sheet.title, row, col)
            # Fallback: return None if merged range not found
            return None

        return self._process_value(cell.value, sheet.title, row, col)

    def read_range(
        self,
        sheet: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
    ) -> list[list[CellValue]]:
        """Read a rectangular range, handling merged cells and #REF! errors.

        Args:
            sheet: The worksheet to read from.
            start_row: 1-based starting row (inclusive).
            end_row: 1-based ending row (inclusive).
            start_col: 1-based starting column (inclusive).
            end_col: 1-based ending column (inclusive).

        Returns:
            A 2D list of cell values.
        """
        result: list[list[CellValue]] = []
        for r in range(start_row, end_row + 1):
            row_data: list[CellValue] = []
            for c in range(start_col, end_col + 1):
                row_data.append(self.read_cell(sheet, r, c))
            result.append(row_data)
        return result

    def get_merged_ranges(self, sheet: Worksheet) -> list[MergedCellRange]:
        """Return all merged cell ranges for boundary detection.

        Args:
            sheet: The worksheet to inspect.

        Returns:
            List of merged cell ranges.
        """
        return list(sheet.merged_cells.ranges)

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _process_value(
        self, value: object, sheet_title: str, row: int, col: int
    ) -> CellValue:
        """Check for #REF! errors and return a clean CellValue."""
        if value is None:
            return None
        if isinstance(value, str) and "#REF!" in value:
            logger.warning(
                "#REF! error in %s!%s%d — replaced with None",
                sheet_title,
                _col_letter(col),
                row,
            )
            return None
        return value  # type: ignore[return-value]


def _col_letter(col: int) -> str:
    """Convert a 1-based column index to an Excel column letter (A, B, …, Z, AA, …)."""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result
